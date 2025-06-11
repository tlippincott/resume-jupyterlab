# functions for resume optimization
import os
import re
import shutil
import urllib.parse
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime

from markdown import markdown
from weasyprint import HTML, CSS

class RandomError(Exception):
    pass

def strip_code_fence(text):
    # remove triple backtick fences (e.g., ```markdown)
    return re.sub(r'```\s*|markdown\s*', '', text, flags=re.IGNORECASE)

def show_pdf(file_name: str) -> str:
    pdf_path = os.path.join("resumes", file_name)
    if os.path.exists(pdf_path):
        return pdf_path
    else:
        return 'resumes/pdf_not_found.pdf'

def create_resume_prompt(resume_string: str, jd_string: str, comp_name_string: str, comp_info_string: str, job_change_bool: bool) -> str:
    """
    Creates a detailed prompt for AI-powered resume section optimization based on a job description, company name, and company information.

    This function generates a structured prompt that guides the AI to:
    - Tailor the resume bullet points and summary to match job requirements
    - Optimize for ATS systems
    - Provide actionable improvement suggestions
    - Format the output in clean Markdown to be utilized in a HTML template

    Args:
        resume_string (str): The input resume bullet points
        jd_string (str): The target job description text
        comp_name_string (str): The company name where applying
        comp_info_string (str): The "About Us" section of the application; information about the company
        job_change_bool (bool): Value to indicate if the summary section should contain a statement about looking for a more customer-focused role (e.g. Helpdesk)

    Returns:
        str: A formatted prompt string containing instructions for resume optimization
    """
    
    return f'''\
You are a professional resume optimization expert specializing in tailoring resume bullet points and summaries to specific job descriptions. Your goal is to optimize my bullet points and provide actionable suggestions for improvement to align with the target role.

Enhance the provided resume bullet points to align with the given job description and create a tailored resume summary. The summary should leverage the enhanced bullet points, the company name, and company information while underscoring how my skills, including 15 years of experience in the technology sector, can benefit the company. The uploaded files will be `resume_string` (a string), `jd_string` (a string), `comp_name_string` (a string), `comp_info_string` (a string), and `job_change_bool` (a boolean).

- Incorporate details from the job description (`jd_string`) to refine the bullet points, ensuring they highlight relevant skills and experiences.
- Craft a resume summary that emphasizes my alignment with the company’s needs and objectives, using the company name (`comp_name_string`) and company information (`comp_info_string`) effectively.

# Steps

1. **Analyze Job Description:** Identify key skills, experiences, and attributes the company is looking for from the `jd_string`.
2. **Modify Bullet Points:** Align each bullet point with these key areas, emphasizing relevant experience and achievements from `resume_string`.
3. **Write Resume Summary:**
   - Begin with a brief introduction summarizing my experience in the technology world.
   - Highlight how specific skills and achievements align with the company's goals.
   - If the Customer Facing Role value (`job_change_bool`) is "True", include a statement about seeking a more customer-focused role. If the Customer Facing Role value (`job_change_bool`) is not "True", do not include a statement about seeking a more customer-focused role.
   - Conclude with the value I can bring to the company.

# Additional Suggestions

- **Additional Skills:** Identify and suggest any additional skills that could enhance my candidacy for roles in the target industry or position.
- **Certifications or Courses:** Recommend certifications or courses that can further establish my expertise and relevance to the target job market.
- **Project Ideas or Experiences:** Propose potential project ideas or experiences that can be pursued to gain or demonstrate relevant expertise.

**Input:**
- Resume in Markdown: {resume_string}
- Job Description: {jd_string}
- Company Name: {comp_name_string}
- Company Info: {comp_info_string}
- Customer Facing Role: {job_change_bool}

# Output Format

- A resume summary consisting of 6-8 sentences, incorporating my experience, skills, and alignment with the company’s objectives. If the Customer Facing Role value (`job_change_bool`) is "True", include a statement about seeking a more customer-focused role. If the Customer Facing Role value (`job_change_bool`) is not "True", do not include a statement about seeking a more customer-focused role. Use information from `comp_name_string` , `comp_info_string`, and `job_change_bool`.
- Enhanced bullet points for each heading (spins, programmer, analyst) that directly relate to the job description (`jd_string`).
- Each bullet point begins with "<li>" and ends with "</li>" and only contains text in the middle.
- Additional Suggestions section detailing actionable points on additional skills, certifications/courses, and project ideas/experiences.
- Return the same updated markdown file as "enhanced-information.md". The file should have a summary, the same number of sections and enhanced bullet points, and additional suggestions placed under the correct headings. Ensure enhancements are made to the "spins", "programmer", and "analyst" sections with enhanced bullet points placed accordingly.

# Examples

## Input
- **Resume In Markdown:**
  - ## summary ##
  - ## spins ##
    - Developed applications using Agile methodologies.
    - Managed a software development team.
  - ## programmer ##
    - Created SQL stored procedures for data pipelines.
    - Was invlolved with every aspect of the software development lifecycle.
  - ## analyst ##
    - Investigated user issues and bugs.
    - Created monthly and weekly data reports.
  - ## Additional Suggestions
- **Job Description:** Seeks project managers with experience in leading technology projects and optimizing team performance.
- **Company Name:** TechVision
- **Company Information:** TechVision is a leading innovator in cloud solutions and AI technologies.
- **Job Change:** True

## Output
## summary ##
With over 15 years in the technology industry, I have honed my ability to lead dynamic teams and drive innovation in fast-paced environments. My experience in managing technology projects aligns seamlessly with TechVision's aim to pioneer cloud solutions and AI advancements. By optimizing team performance and fostering collaboration, I have consistently delivered top-tier results. Additionally, I am seeking a role that is more customer-focused to leverage my strategic insights in enhancing client engagement. At TechVision, I plan to leverage these skills to spearhead projects that fuel growth and differentiation in the market. My strategic perspective and commitment to quality will support your mission of innovation. I am excited about the opportunity to contribute to TechVision's success and drive future achievements.

## spins ##
<li>Successfully led cross-functional teams in developing key applications, utilizing Agile methodologies to enhance productivity and product quality.</li>
<li>Managed a high-performing software development team, achieving a 20% increase in project delivery efficiency</li>

## programmer ##
<li>Created complex SQL stored procedures utilized in nightly batch processing</li>
<li>Heavily involved in all aspects of the development lifecycle, from requirements gathering to user testing and training</li>

## analyst ##
<li>Was first line of contact for user issues, including installation problems and software bugs</li>
<li>Created and automated complex monthly and weekly reports, allowing management to make decisions based on data analysis</li>

## Additional Suggestions
**Additional Skills:** Leadership in collaborative project environments.
**Certifications or Courses:** Agile Project Management Certification.
**Project Ideas or Experiences:** Develop an AI-focused project that aligns with cloud services.

# Notes

- Ensure that the enhanced bullet points and summary narrative maintain a formal and professional tone.
- Customize the summary to reflect a strong understanding of the company’s mission and values.
'''

def create_cover_letter_prompt(bullet_point_string: str, jt_string: str, jd_string: str, comp_name_string: str, comp_info_string: str, job_change_bool: bool) -> str:
    """
    Creates a detailed prompt for AI-powered cover letter creation based on resume bullet points, job title, job description, company name, and company information.

    This function generates a structured prompt that guides the AI to:
    - Create the body of a cover letter for a specific job posting
    - Format the output in clean Markdown to be utilized in a HTML template

    Args:
        bullet_point_string (str): The enhanced resume bullet points created previously by ChatGPT
        jt_string (str): The job title
        jd_string (str): The target job description text
        comp_name_string (str): The company name where applying
        comp_info_string (str): The "About Us" section of the application; information about the company
        job_change_bool (bool): Value to indicate if the cover letter text should contain a statement about looking for a more customer-focused role (e.g. Help Desk)

    Returns:
        str: A formatted prompt string containing instructions for creating the body of a cover letter
    """
    
    return f'''\
You are a professional cover letter expert specializing in tailoring cover letters to specific job postings. Your goal is to write a professional, concise, and effective body for a cover letter that aligns with the provided job opportunity.

The cover letter should leverage the resume bullet points, job title, company name, company information, and job description, while underscoring how my skills, including 15 years of experience in the technology sector, can benefit the company. If the 'job_change_bool' value is "True," also include a statement that I am looking for a role that is more customer-focused. The uploaded files will be Resume Bullet Points (a string), Job Title (a string), Company Name (a string), Company Information (a string), Job Description (a string), Customer Facing Role (a boolean). The total word count of the cover letter should be 280-350 words.

**Input:**
- Resume Bullet Points: {bullet_point_string}
- Job Title: {jt_string}
- Company Name: {comp_name_string}
- Company Information: {comp_info_string}
- Job Description: {jd_string}
- Customer Facing Role: {job_change_bool}

# Steps

1. **Analyze the Information** 
   - Review the resume bullet points from the Resume Bullet Points.
   - Understand the job title from the Job Title, company name from the Company Name, and company information from the Company Information.
   - Thoroughly read the job description from the Job Description to identify key responsibilities and qualifications.

2. **Draft the Cover Letter**
   - Clearly state the job title and company I am applying to in the opening sentence.
   - Use the provided resume bullet points to highlight my relevant experience and achievements.
   - Incorporate company-specific information and align my skills with what the company is seeking based on the job description.
   - Conclude with a strong closing paragraph expressing enthusiasm and inviting further communication.
   
3. **Ensure the content is concise, well-structured, and free of errors.**

# Output Format

Generate the cover letter as a markdown text file named "generated_cover_letter.md". Ensure it includes:

- A structured body that logically presents my qualifications and connection to the job role.
- A closing remark.
- Contact information is not necessary.
- The wording "## cover_letter_body ##" and the remaining text underneath, with each paragraph beinning with a "<p>" and ending with a "</p>".
- The total word count should be 280-350 words.

# Notes

- Focus on creating a positive tone and demonstrating value to the company.
- Tailor the language and examples based on the job description and company context provided.
- Avoid generic statements; personalize content to the role and company.

# Example 

**Input:**
- Resume Bullet Points:
  - Assisted end users with software installation and setup.
  - Created user manuals for publication on the company web site.
- Job Title: Help Desk Associate
- Company Name: TechVision
- Company Information: TechVision is a leading innovator in cloud solutions and AI technologies.
- Job Description: Seeks help desk associates to assist customers with problem resolution and to answer questions in an efficient and courteous manner.
- Customer Facing Role: True

**Output:**
## cover_letter_body ##
<p>I am writing to express my interest in the Help Desk Associate position at TechVision. With 15 years of experience in the technology field and proven success in customer assistance, I am excited about the opportunity to contribute to TechVision with my skills in software installation and setup, answering customer questions, and creating user manuals.</p>

<p>Throughout my career, I have been heavily involved in all aspects of the software development lifecycle. I am very comfortable bridging the gap between the technical and the non-technical.</p>

<p>TechVision stands out to me because of their pioneering cloud solutions and AI advancements, and I am eager to bring my expertise in customer satisfaction to your esteemed company.</p>

<p>Thank you for considering my application. I look forward to the possibility of discussing how I can contribute to the success of TechVison.</p>
'''
    
def get_response(prompt: str, my_api_key: str, model: str = "gpt-4o-mini", temperature: float = 0.7) -> str:
    """
    Sends a prompt to OpenAI's API and returns the requested output.

    This function:
    - Initializes the OpenAI client
    - Makes an API call with the provided prompt
    - Returns the generated response

    Args:
        prompt (str): The formatted prompt containing the required information
        api_key (str): OpenAI API key for authentication
        model (str, optional): The OpenAI model to use. Defaults to "gpt-4-turbo-preview"
        temperature (float, optional): Controls randomness in the response. Defaults to 0.7

    Returns:
        str: The AI-generated response

    Raises:
        OpenAIError: If there's an issue with the API call
    """
    
    # Setup API client
    client = OpenAI(api_key=my_api_key)

    # Make API call
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "Expert resume writer"},
            {"role": "user", "content": prompt}
        ],
        temperature=temperature
    )

    # Extract and return response
    return response.choices[0].message.content

def process_resume(resume, jd_string, comp_name_string, comp_info_string, job_change):
    """
    Process resume sections against a job description, a company name, and company information, to create optimized bullet points and a summary.

    Args:
        resume (file): A file object containing the bullet points in markdown format
        jd_string (str): The job description text to optimize the resume against
        comp_name_string (str): The company name where applying to use in the resume summary section
        comp_info_string (str): The "About Us" section of the application, also to be utilized in creating the resume summary section
        job_change_bool (bool): Should the summary section include the statement you are looking for a job change which is more customer-focused (e.g. Helpdesk Position)

    Returns:
        tuple: A tuple containing three elements:
            - str: The optimized resume sections in markdown format (for display)
            - str: The same optimized resume sections (for editing)
            - str: Suggestions for improving the resume
    """
    
    # read resume
    with open(resume, "r", encoding="utf-8") as file:
        resume_string = file.read()

    # set job change boolean value based on radio button selection
    if job_change == 'Yes':
        job_change_bool:bool = True
    else:
        job_change_bool:bool = False

    # create prompt
    prompt = create_resume_prompt(resume_string, jd_string, comp_name_string, comp_info_string, job_change_bool)

    # load the .env file
    load_dotenv()

    # assign the ChatGPT api key
    my_api_key = os.getenv("API_KEY")

    # generate response
    response_string = get_response(prompt, my_api_key)
    response_list = response_string.split("## Additional Suggestions")
    
    # extract new resume and suggestions for improvement
    new_resume_sections = response_list[0]
    suggestions = "## Additional Suggestions \n\n" + response_list[1]

    # remove "```markdown" from returned file
    clean_resume_sections = strip_code_fence(new_resume_sections)

    # save the clean resume sections
    with open("resumes/resume_new.md", 'w') as f:
        f.write(clean_resume_sections)

    return clean_resume_sections, suggestions

def split_bullet_points():
    # read the file
    with open('resumes/resume_new.md', 'r', encoding='utf-8') as f:
        content = f.read()
    
    # split summary and bullet points
    split_marker = '## spins ##'
    sections = content.split(split_marker)
    
    if len(sections) > 1:
        # include the split marker at the beginning of the second part
        enhanced_bullet_points = split_marker + sections[1]
    
        # save the enhanced bullet points to a new file
        with open('resumes/enhanced_bullet_points.md', 'w', encoding='utf-8') as f:
            f.write(enhanced_bullet_points)
    else:
        raise RandomError("'resume_new.md' not split as expected")

def process_cover_letter(jt_string, jd_string, comp_name_string, comp_info_string, job_change):
    """
    Process resume bullet points and job posting information in order to create the body of a cover letter.

    Args:
        bullet_points (file): A file object containing the enhanced bullet points in markdown format
        jt_string (str): The job title in the posting
        jd_string (str): The job description text
        comp_name_string (str): The company name where applying
        comp_info_string (str): The "About Us" section of the application
        job_change_bool (bool): Should the cover letter body include the statement you are looking for a job change which is more customer-focused (e.g. Help Desk Associate)

    Returns:
        - str: The body of the cover letter
    """
    
    # separate the summary from the enhanced bullet points
    try:
        split_bullet_points()
    except NotFoundError as e:
        return "'resume_new.md' not split as expected"
        
    # read enhanced bullet point file
    with open('resumes/enhanced_bullet_points.md', "r", encoding="utf-8") as file:
        bullet_point_string = file.read()

    # set job change boolean value based on radio button selection
    if job_change == 'Yes':
        job_change_bool:bool = True
    else:
        job_change_bool:bool = False

    # create prompt
    prompt = create_cover_letter_prompt(bullet_point_string, jt_string, jd_string, comp_name_string, comp_info_string, job_change_bool)

    # load the .env file
    load_dotenv()

    # assign the ChatGPT api key
    my_api_key = os.getenv("API_KEY")

    # generate response
    cover_letter_string = get_response(prompt, my_api_key)

    # remove "```" from returned file
    clean_cover_letter_body = strip_code_fence(cover_letter_string)

    # save the clean resume sections
    with open("resumes/cover_letter_new.md", 'w') as f:
        f.write(clean_cover_letter_body)

    return clean_cover_letter_body, "Successfully generated cover letter"

def save_edits(section_edits):
    """
    Save resume section edits as a Markdown file

    Args:
        section_edits (str): Final resume sections after any edits

    Returns:
        str: A message indicating the save was a success/failure
    """

    try:
        with open("resumes/resume_new.md", 'w') as f:
            f.write(section_edits)

        return "Edit section saved successfully!"
    except Exception as e:
        return f"Failed to save edits: {str(e)}"

def save_cover_edits(cover_body_edits):
    """
    Save cover letter body edits as a Markdown file

    Args:
        cover_body_edits (str): Final cover letter body section after any edits

    Returns:
        str: A message indicating the save was a success/failure
    """

    try:
        with open("resumes/cover_letter_new.md", 'w') as f:
            f.write(cover_body_edits)

        return "Cover letter edits saved successfully!"
    except Exception as e:
        return f"Failed to save edits: {str(e)}"

def extract_section(sections, md_section_file):
    result = {}
    for section in sections:
        # regex to match heading and capture text until next heading or end of file
        pattern = rf'##\s*{section}\s*##\s*\n(.*?)(?=\n##|\Z)'
        match = re.search(pattern, md_section_file, re.DOTALL | re.MULTILINE)
        
        if match:
            extracted = match.group(1).strip()
        else:
            extracted = None
        
        result[section] = extracted
    return result

def construct_html(md_file, html_template_file, md_file_sections):
    # extract Markdown sections
    with open(md_file, 'r', encoding='utf-8') as file:
        md_section_file = file.read()

    md_sections = extract_section(md_file_sections, md_section_file)
    
    # get the HTML template file
    with open(html_template_file, 'r', encoding='utf-8') as file:
        html_template = file.read()

    # insert the corresponding sections into the HTML template
    final_html = html_template.format(**md_sections)

    return final_html

def export_resume():
    """
    Save the enhanced resume as a PDF file
    """

    md_file = 'resumes/resume_new.md'
    html_template_file = 'resumes/resume.html'
    resume_sections = ('summary', 'spins', 'programmer', 'analyst')
    
    try:
        # create HTML file
        final_html_resume = construct_html(md_file, html_template_file, resume_sections)
        
        # convert HTML to PDF and save
        HTML(string=final_html_resume).write_pdf('resumes/Terry_Lippincott_Resume_2025.pdf', stylesheets=['resumes/resume_style.css'])

        return f"Successfully generated resume 'Terry_Lippincott_Resume_2025.pdf'"
    except Exception as e:
        return f"Failed to generate resume: {str(e)}"

def export_cover_letter():
    """
    Save the cover letter as a PDF file
    """

    md_file = 'resumes/cover_letter_new.md'
    html_template_file = 'resumes/cover_letter.html'
    cover_letter_sections = ('cover_letter_body',)
    
    try:
        # create HTML file
        final_html_cover_letter = construct_html(md_file, html_template_file, cover_letter_sections)
        
        # convert HTML to PDF and save
        HTML(string=final_html_cover_letter).write_pdf('resumes/Terry_Lippincott_Cover_2025.pdf', stylesheets=['resumes/cover_style.css'])

        return f"Successfully generated cover letter 'Terry_Lippincott_Cover_2025.pdf'"
    except Exception as e:
        return f"Failed to generate cover letter: {str(e)}"

def move_to_downloads(file_name: str) -> str:
    """
    Move resume pdf and cover letter pdf to the Downloads folder

    Args:
        str: Name of file to be moved

    Returns:
        str: A message indicating the move was a success/failure
    """

    try:
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")

        # move resume pdf
        source_file = os.path.join("resumes", file_name)
        destination = os.path.join(downloads_dir, os.path.basename(source_file))
        shutil.move(source_file, destination)

        return f"{file_name} moved successfully!"
    except Exception as e:
        return f"Failed to move pdf files: {str(e)}"
  
def save_job_posting_info(company_name: str, job_title: str, job_posting_site: str):
    # define file name
    file_name = company_name + " " + job_title + ".txt"

    # construct file path for saving/moving job posting text file
    destination_folder = os.path.join(os.path.expanduser("~"), "Library/Group Containers/UBF8T346G9.Office")
    destination_file = os.path.join(destination_folder, file_name)

    # job posting text file location
    link_file_url = f"file://{destination_file}"

    # Excel hyperlink
    excel_job_text_link = f'=HYPERLINK("{link_file_url}", "{file_name}")'

    # timestamp
    today = datetime.today().strftime("%m/%d/%Y")

    # job posting data to append to Excel worksheet
    new_row = ["Waiting", company_name, job_title, excel_job_text_link, today, "Yes", today, job_posting_site]

    # Excel file path
    excel_file_path = os.path.join(os.path.expanduser("~"), "Desktop", "Job_Search_Tracker.xlsm")

    try:
        if os.path.exists(excel_file_path):
            wb = load_workbook(excel_file_path, keep_vba=True)
            ws = wb["My Progress"]
    
            # find first blank row in column A
            row_num = 1
            while ws[f"A{row_num}"].value is not None:
                row_num += 1
    
            # write row
            for col_index, value in enumerate(new_row, start=1):
                ws.cell(row=row_num, column=col_index, value=value)
    
            wb.save(excel_file_path)
    
        # move and rename the job posting text file
        downloads_folder = os.path.expanduser("~/Downloads")
        source_file = os.path.join(downloads_folder, "web_page_text.txt")
        shutil.move(source_file, destination_file)

        return "Job posting information added successfully!"
    
    except Exception as e:
        return f"Failed to add job posting information: {str(e)}"
